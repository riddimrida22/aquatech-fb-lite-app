#!/usr/bin/env python3
from __future__ import annotations

import csv
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
WORKSHEET_CSV = ROOT / "docs" / "compliance" / "2025-tax-1120s-draft-worksheet.csv"
CATEGORY_MAP_CSV = ROOT / "docs" / "compliance" / "2025-tax-pl-category-map.csv"
OUTPUT_XLSX = ROOT / "docs" / "compliance" / "2024-vs-2025-tax-line-comparison.xlsx"
DOWNLOADS = Path("/mnt/c/Users/bertr/Downloads")
TAX_FOLDER = DOWNLOADS / "Aquatech-Tax-Documents"
PERSONAL_CARD_EXPENSE_ADJUSTMENT = 25000.0


FORM_LINE_ORDER = [
    "1a",
    "1b",
    "1c",
    "2",
    "3",
    "4",
    "5",
    "6",
    "7",
    "8",
    "9",
    "10",
    "11",
    "12",
    "13",
    "14",
    "15",
    "16",
    "17",
    "18",
    "19",
    "20",
    "21",
    "22",
]


LINE_2024_DATA = [
    {"form_line": "1a", "label": "Gross receipts or sales", "amount": 216919.00, "source": "Direct PDF text extraction", "note": "Recovered from the flattened 2024 Form 1120-S page-1 amount sequence."},
    {"form_line": "1b", "label": "Less returns and allowances", "amount": 0.00, "source": "Derived", "note": "No separate returns-and-allowances amount appeared in the flattened PDF sequence."},
    {"form_line": "1c", "label": "Balance", "amount": 216919.00, "source": "Direct PDF text extraction", "note": "Recovered from the flattened 2024 Form 1120-S page-1 amount sequence."},
    {"form_line": "2", "label": "Cost of goods sold", "amount": 0.00, "source": "Derived", "note": "Inferred because line 1c and gross profit match and no COGS amount appeared in the flattened PDF sequence."},
    {"form_line": "3", "label": "Gross profit", "amount": 216919.00, "source": "Direct PDF text extraction", "note": "Recovered from the flattened 2024 Form 1120-S page-1 amount sequence."},
    {"form_line": "4", "label": "Net gain/loss from Form 4797", "amount": 0.00, "source": "Derived", "note": "No Form 4797 amount appeared in the provided 2024 return PDF text."},
    {"form_line": "5", "label": "Other income/loss", "amount": 0.00, "source": "Derived", "note": "No page-1 other-income amount appeared in the provided 2024 return PDF text."},
    {"form_line": "6", "label": "Total income/loss", "amount": 216919.00, "source": "Direct PDF text extraction", "note": "Also supported by Form 8453-CORP page 2, which shows total income (loss) from Form 1120-S line 6."},
    {"form_line": "7", "label": "Compensation of officers", "amount": 30414.00, "source": "Inferred from line-order sequence", "note": "Mapped from the nonzero page-1 amount sequence in the flattened PDF after line 6."},
    {"form_line": "8", "label": "Salaries and wages", "amount": 57780.00, "source": "Inferred from line-order sequence", "note": "Mapped from the nonzero page-1 amount sequence in the flattened PDF after line 6."},
    {"form_line": "9", "label": "Repairs and maintenance", "amount": 0.00, "source": "Derived", "note": "No nonzero repairs amount appeared in the flattened 2024 page-1 amount sequence."},
    {"form_line": "10", "label": "Bad debts", "amount": 0.00, "source": "Derived", "note": "No nonzero bad-debt amount appeared in the flattened 2024 page-1 amount sequence."},
    {"form_line": "11", "label": "Rents", "amount": 8500.00, "source": "Inferred from line-order sequence", "note": "Mapped from the nonzero page-1 amount sequence in the flattened PDF after lines 7-8."},
    {"form_line": "12", "label": "Taxes and licenses", "amount": 12600.00, "source": "Inferred from line-order sequence", "note": "Mapped from the nonzero page-1 amount sequence in the flattened PDF after lines 7-11."},
    {"form_line": "13", "label": "Interest", "amount": 10171.00, "source": "Inferred from line-order sequence", "note": "Mapped from the nonzero page-1 amount sequence in the flattened PDF after lines 7-12."},
    {"form_line": "14", "label": "Depreciation", "amount": 323.00, "source": "Direct PDF text extraction", "note": "Supported by the Form 4562 page in the 2024 return PDF, which shows total depreciation of 323."},
    {"form_line": "15", "label": "Depletion", "amount": 0.00, "source": "Derived", "note": "No depletion amount appeared in the provided 2024 return PDF text."},
    {"form_line": "16", "label": "Advertising", "amount": 2500.00, "source": "Inferred from line-order sequence", "note": "Mapped from the nonzero page-1 amount sequence in the flattened PDF after depreciation."},
    {"form_line": "17", "label": "Pension, profit-sharing, etc. plans", "amount": 13788.00, "source": "Inferred from line-order sequence", "note": "Mapped from the nonzero page-1 amount sequence in the flattened PDF after advertising."},
    {"form_line": "18", "label": "Employee benefit programs", "amount": 5475.00, "source": "Inferred from line-order sequence", "note": "Mapped from the nonzero page-1 amount sequence in the flattened PDF after line 17."},
    {"form_line": "19", "label": "Energy efficient buildings deduction", "amount": 0.00, "source": "Derived", "note": "No Form 7205 / energy-efficient-building deduction amount appeared in the 2024 return PDF text."},
    {"form_line": "20", "label": "Other deductions (attach statement)", "amount": 123506.00, "source": "Direct PDF text extraction", "note": "Supported directly by the 2024 `Other Deductions - Supporting Details for Form 1120-S, Line 20` statement."},
    {"form_line": "21", "label": "Total deductions", "amount": 265057.00, "source": "Direct PDF text extraction", "note": "Recovered from the flattened 2024 page-1 amount sequence and ties to line 6 less line 22."},
    {"form_line": "22", "label": "Ordinary business income/loss", "amount": -48138.00, "source": "Direct PDF text extraction", "note": "Supported by the 2024 Schedule K-1 and page-1 amount sequence."},
]


OTHER_DEDUCTIONS_2024 = [
    ("Accounting fees", 800.00, [("Professional Services", "Accounting")], "Direct 2024 support statement line item."),
    ("Bank fees", 500.00, [("Other Expenses", "Bank Fees")], "Direct 2024 support statement line item."),
    ("Insurance", 6500.00, [("Other Expenses", "Business Insurance")], "Direct 2024 support statement line item."),
    ("Legal and professional fees", 2500.00, [("Professional Services", "Legal Fees"), ("Professional Services", "Professional Services (general)")], "Approximate 2025 mapping to non-accounting professional-service categories."),
    ("Professional dues and subscriptions", 250.00, [], "No clean one-to-one 2025 category was identified from the provided books."),
    ("Supplies", 750.00, [("Office Expenses & Postage", "Office Expenses & Postage (general)"), ("Office Expenses & Postage", "Stationery")], "Approximate 2025 mapping to general office-expense supplies categories."),
    ("Telephone", 2400.00, [("Utilities", "Phone")], "Direct 2025 phone utility mapping."),
    ("Deductible non-entertainment meals exp. subject to limits", 1750.00, [("Meals & Entertainment", "Meals & Entertainment (general)"), ("Meals & Entertainment", "Restaurants/Dining")], "2025 books do not separately break out 50%-limit meals versus fully deductible meals."),
    ("Travel and non-entertainment meals expense not subject to limits", 12500.00, [("Travel", "Airfare"), ("Travel", "Hotel/Lodging/Accommodation"), ("Travel", "Taxi & Parking"), ("Travel", "Travel (general)")], "Approximate 2025 travel-category total."),
    ("Mileage", 5000.00, [("Car & Truck Expenses", "Mileage")], "Direct mileage-category comparison."),
    ("Business development", 90556.00, [("Professional Services", "Business Development")], "2025 books show a much smaller explicitly tagged business-development amount."),
]


def load_2025_lines(path: Path) -> dict[str, dict[str, object]]:
    rows: dict[str, dict[str, object]] = {}
    with path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            if row["Section"] != "Page 1":
                continue
            rows[row["Form Line"]] = {
                "label": row["Label"],
                "cash": float(row["Cash Basis Amount"]),
                "accrual": float(row["Accrual Basis Amount"]),
                "note": row["Note"],
            }
    return rows


def load_2025_category_map(path: Path) -> dict[tuple[str, str], tuple[float, float]]:
    rows: dict[tuple[str, str], tuple[float, float]] = {}
    with path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            rows[(row["Category"], row["Subcategory"])] = (float(row["Cash Total"]), float(row["Accrual Total"]))
    return rows


def adjusted_amount(form_line: str, amount: float) -> float:
    if form_line in {"20", "21"}:
        return amount + PERSONAL_CARD_EXPENSE_ADJUSTMENT
    if form_line == "22":
        return amount - PERSONAL_CARD_EXPENSE_ADJUSTMENT
    return amount


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 12), 48)


def currency(value: float) -> float:
    return round(value, 2)


def build_workbook() -> list[Path]:
    lines_2025 = load_2025_lines(WORKSHEET_CSV)
    category_map = load_2025_category_map(CATEGORY_MAP_CSV)

    wb = Workbook()
    ws = wb.active
    ws.title = "1120S Line Compare"
    ws.freeze_panes = "A2"

    header = [
        "Form Line",
        "Label",
        "2024 Filed Amount",
        "2024 Source",
        "2024 Note",
        "2025 Cash Books",
        "2025 Cash + $25k Adj",
        "Cash Delta vs 2024",
        "2025 Accrual Books",
        "2025 Accrual + $25k Adj",
        "Accrual Delta vs 2024",
        "2025 Note",
    ]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    line_2024_by_form = {row["form_line"]: row for row in LINE_2024_DATA}
    for form_line in FORM_LINE_ORDER:
        prior = line_2024_by_form[form_line]
        current = lines_2025.get(form_line, {})
        cash = float(current.get("cash", 0.0))
        accrual = float(current.get("accrual", 0.0))
        cash_adj = adjusted_amount(form_line, cash)
        accrual_adj = adjusted_amount(form_line, accrual)
        prior_amount = float(prior["amount"])
        ws.append(
            [
                form_line,
                prior["label"],
                currency(prior_amount),
                prior["source"],
                prior["note"],
                currency(cash),
                currency(cash_adj),
                currency(cash_adj - prior_amount),
                currency(accrual),
                currency(accrual_adj),
                currency(accrual_adj - prior_amount),
                current.get("note", ""),
            ]
        )

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=10):
        for cell in row:
            cell.number_format = '#,##0.00;[Red]-#,##0.00'

    ws2 = wb.create_sheet("2024 Other Deductions")
    ws2.freeze_panes = "A2"
    ws2.append(
        [
            "2024 Detail Line",
            "2024 Amount",
            "2025 Cash Candidate",
            "2025 Accrual Candidate",
            "2025 Mapping Used",
            "Note",
        ]
    )
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    total_2024_other = 0.0
    total_2025_other_cash = 0.0
    total_2025_other_accrual = 0.0
    for label, amount_2024, mappings, note in OTHER_DEDUCTIONS_2024:
        cash_total = sum(category_map.get(key, (0.0, 0.0))[0] for key in mappings)
        accrual_total = sum(category_map.get(key, (0.0, 0.0))[1] for key in mappings)
        mapping_text = ", ".join(f"{category} / {subcategory}" for category, subcategory in mappings) or "No direct mapping selected"
        ws2.append([label, currency(amount_2024), currency(cash_total), currency(accrual_total), mapping_text, note])
        total_2024_other += amount_2024
        total_2025_other_cash += cash_total
        total_2025_other_accrual += accrual_total

    ws2.append(
        [
            "Total of listed 2024 support detail",
            currency(total_2024_other),
            currency(total_2025_other_cash),
            currency(total_2025_other_accrual),
            "",
            "This total covers the extracted 2024 Line 20 support statement categories shown in the PDF.",
        ]
    )
    for row in ws2.iter_rows(min_row=2, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = '#,##0.00;[Red]-#,##0.00'

    ws3 = wb.create_sheet("Assumptions")
    ws3.append(["Item", "Detail"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    assumptions = [
        (
            "2024 page-1 values",
            "Lines 1a, 1c, 3, 6, 20, 21, and 22 were recoverable from flattened PDF text or direct support statements. Lines 7, 8, 11, 12, 13, 16, 17, and 18 were inferred from the ordered nonzero amount sequence on the flattened 2024 Form 1120-S page 1.",
        ),
        (
            "2024 depreciation",
            "Line 14 depreciation is supported by the 2024 Form 4562 page, which shows total depreciation of 323.",
        ),
        (
            "2025 values",
            "2025 values come from docs/compliance/2025-tax-1120s-draft-worksheet.csv generated from the current AquatechPM tax packet workflow.",
        ),
        (
            "Personal-card adjustment",
            "The workbook shows both book values and values adjusted for the additional $25,000 of shareholder-paid business expenses. That adjustment only changes lines 20, 21, and 22.",
        ),
        (
            "Use limitation",
            "This workbook is a working-paper comparison and not a substitute for confirming the exact filed 2024 return boxes and amounts from the original tax software records.",
        ),
    ]
    for item, detail in assumptions:
        ws3.append([item, detail])

    autosize_columns(ws)
    autosize_columns(ws2)
    autosize_columns(ws3)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)

    copied = []
    for destination_dir in [DOWNLOADS, TAX_FOLDER]:
        destination_dir.mkdir(parents=True, exist_ok=True)
        destination = destination_dir / OUTPUT_XLSX.name
        shutil.copy2(OUTPUT_XLSX, destination)
        copied.append(destination)

    return [OUTPUT_XLSX, *copied]


def main() -> None:
    for path in build_workbook():
        print(f"Wrote: {path}")


if __name__ == "__main__":
    main()
