#!/usr/bin/env python3
from __future__ import annotations

import csv
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
COMPLIANCE_DIR = ROOT / "docs" / "compliance"
WORKSHEET_CSV = COMPLIANCE_DIR / "2025-tax-1120s-draft-worksheet.csv"
MANUAL_ADJUSTMENTS_CSV = COMPLIANCE_DIR / "2025-tax-manual-adjustments.csv"
SHAREHOLDER_MATCH_CSV = COMPLIANCE_DIR / "2025-tax-shareholder-loan-match.csv"
CHECKLIST_MD = COMPLIANCE_DIR / "2025-tax-cpa-handoff-checklist.md"
OUTPUT_XLSX = COMPLIANCE_DIR / "2025-draft-1120S-NOT-FOR-FILING.xlsx"
OUTPUT_MD = COMPLIANCE_DIR / "2025-draft-1120S-NOT-FOR-FILING.md"
DOWNLOADS = Path("/mnt/c/Users/bertr/Downloads")
TAX_FOLDER = DOWNLOADS / "Aquatech-Tax-Documents"

ENTITY_INFO = [
    ("Legal name", "Aquatech Engineering PC", "From 2024 filed return PDF", "Verify unchanged for the 2025 filing."),
    ("EIN", "46-1465989", "From 2024 filed return PDF", ""),
    ("Address", "15 Bonita Vista Road, Mount Vernon, NY 10552", "From 2024 filed return PDF", "Verify if principal office changed in 2025."),
    ("Business activity code", "541330", "From 2024 filed return PDF", "Engineering services / consulting."),
    ("Business activity", "Engineering consulting", "From 2024 Form 4562 text", ""),
    ("Date incorporated", "2012-11-20", "From 2024 filed return PDF", ""),
    ("S election effective date", "2013-01-01", "From flattened 2024 return text", "Verify against original election records."),
    ("Tax year", "2025-01-01 through 2025-12-31", "Current packet assumption", "Calendar-year return."),
    ("Form due date", "2026-03-16", "IRS 2025 Form 1120-S instructions", "March 15, 2026 falls on Sunday."),
    ("Extension form", "Form 7004", "IRS 2025 Form 1120-S instructions", "Must be filed by March 16, 2026 if extending."),
    ("2024 filed accounting method", "Unknown from provided PDF text", "Needs confirmation", "Do not choose 2025 cash vs accrual until the filed 2024 method is confirmed."),
    ("2025 shareholder assumption", "Bertrand A. Byrne, 100% shareholder", "Based on 2024 Schedule K-1", "Verify if any ownership change occurred in 2025."),
]

EXTRA_PAGE_1_LINES = [
    ("23a", "Excess net passive income or LIFO recapture tax"),
    ("23b", "Tax from Schedule D (Form 1120-S)"),
    ("23c", "Other taxes and penalties"),
    ("24a", "Current year's estimated tax payments and prior-year overpayment"),
    ("24b", "Tax deposited with Form 7004"),
    ("24c", "Credit for federal tax paid on fuels"),
    ("24d", "Elective payment election amount from Form 3800"),
    ("24z", "Total payments"),
    ("25", "Estimated tax penalty"),
    ("26", "Amount owed"),
    ("27", "Overpayment"),
    ("28a", "Credited to 2026 estimated tax"),
    ("28b", "Refunded"),
    ("28c", "Routing number"),
    ("28d", "Account number"),
    ("28e", "Type of account"),
]

PAGE_1_LINE_ORDER = {
    "1a": 1,
    "1b": 2,
    "1c": 3,
    "2": 4,
    "3": 5,
    "4": 6,
    "5": 7,
    "6": 8,
    "7": 9,
    "8": 10,
    "9": 11,
    "10": 12,
    "11": 13,
    "12": 14,
    "13": 15,
    "14": 16,
    "15": 17,
    "16": 18,
    "17": 19,
    "18": 20,
    "19": 21,
    "20": 22,
    "21": 23,
    "22": 24,
    "23a": 25,
    "23b": 26,
    "23c": 27,
    "24a": 28,
    "24b": 29,
    "24c": 30,
    "24d": 31,
    "24z": 32,
    "25": 33,
    "26": 34,
    "27": 35,
    "28a": 36,
    "28b": 37,
    "28c": 38,
    "28d": 39,
    "28e": 40,
}

SCHEDULE_K_LINES = [
    ("1", "Ordinary business income (loss)"),
    ("2", "Net rental real estate income (loss)"),
    ("3c", "Other net rental income (loss)"),
    ("4", "Interest income"),
    ("5a", "Ordinary dividends"),
    ("5b", "Qualified dividends"),
    ("6", "Royalties"),
    ("7", "Net short-term capital gain (loss)"),
    ("8a", "Net long-term capital gain (loss)"),
    ("8b", "Collectibles (28%) gain (loss)"),
    ("8c", "Unrecaptured section 1250 gain"),
    ("9", "Net section 1231 gain (loss)"),
    ("10", "Other income (loss)"),
    ("11", "Section 179 deduction"),
    ("12a", "Cash charitable contributions"),
    ("12b", "Noncash charitable contributions"),
    ("12c", "Investment interest expense"),
    ("12d", "Section 59(e)(2) expenditures"),
    ("12e", "Other deductions"),
    ("16a", "Tax-exempt interest income"),
    ("16b", "Other tax-exempt income"),
    ("16c", "Nondeductible expenses"),
    ("16d", "Distributions"),
    ("16e", "Repayment of loans from shareholders"),
    ("16f", "Foreign taxes paid or accrued"),
]


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 12), 52)


def money(value: float) -> float:
    return round(value, 2)


def load_page_1_lines(path: Path) -> dict[str, dict[str, object]]:
    data: dict[str, dict[str, object]] = {}
    with path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            if row["Section"] != "Page 1":
                continue
            data[row["Form Line"]] = {
                "label": row["Label"],
                "cash": float(row["Cash Basis Amount"]),
                "accrual": float(row["Accrual Basis Amount"]),
                "note": row["Note"],
            }
    return data


def load_manual_adjustments(path: Path) -> tuple[float, float]:
    loan_repayment = 0.0
    personal_card_expense = 0.0
    with path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            label = row["Label"].lower()
            amount = float(row["Amount"])
            if "shareholder loan" in label:
                loan_repayment += amount
            if "personal credit card" in label:
                personal_card_expense += amount
    return loan_repayment, personal_card_expense


def load_shareholder_match(path: Path) -> dict[str, str]:
    metrics: dict[str, str] = {}
    with path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            metrics[row["Metric"]] = row["Amount"]
    return metrics


def load_checklist_items(path: Path) -> tuple[list[str], list[str]]:
    missing: list[str] = []
    review: list[str] = []
    if not path.exists():
        return missing, review
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if line.startswith("- MISSING:"):
            missing.append(line.removeprefix("- MISSING:").strip())
        elif line.startswith("- REVIEW:"):
            review.append(line.removeprefix("- REVIEW:").strip())
    return missing, review


def build_schedule_k_rows(page_1_lines: dict[str, dict[str, object]], loan_repayment: float, personal_card_expense: float) -> list[list[object]]:
    ordinary_cash = float(page_1_lines["22"]["cash"])
    ordinary_accrual = float(page_1_lines["22"]["accrual"])
    rows: list[list[object]] = []
    for line_code, label in SCHEDULE_K_LINES:
        cash_value: float | None = None
        cash_adj_value: float | None = None
        accrual_value: float | None = None
        accrual_adj_value: float | None = None
        status = "Blank pending support"
        note = "No reliable 2025 amount identified from the current packet."
        if line_code == "1":
            cash_value = ordinary_cash
            cash_adj_value = ordinary_cash - personal_card_expense
            accrual_value = ordinary_accrual
            accrual_adj_value = ordinary_accrual - personal_card_expense
            status = "Draft value from current packet"
            note = "Pulled from Form 1120-S page 1 line 22 in the current 2025 draft worksheet."
        elif line_code == "16e":
            cash_value = loan_repayment
            cash_adj_value = loan_repayment
            accrual_value = loan_repayment
            accrual_adj_value = loan_repayment
            status = "Manual adjustment"
            note = "User-directed shareholder loan repayment carried to Schedule K informationally."
        elif line_code == "16d":
            status = "Unresolved owner activity"
            note = "Do not auto-fill distributions while shareholder transfer activity is still under review."

        rows.append([line_code, label, cash_value, cash_adj_value, accrual_value, accrual_adj_value, status, note])
    return rows


def build_outputs() -> list[Path]:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    page_1_lines = load_page_1_lines(WORKSHEET_CSV)
    loan_repayment, personal_card_expense = load_manual_adjustments(MANUAL_ADJUSTMENTS_CSV)
    shareholder_match = load_shareholder_match(SHAREHOLDER_MATCH_CSV)
    missing_items, review_items = load_checklist_items(CHECKLIST_MD)

    wb = Workbook()
    warning_fill = PatternFill("solid", fgColor="FDE9D9")
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    ws = wb.active
    ws.title = "Read Me"
    ws.append(["2025 Draft Form 1120-S Packet", "DRAFT NOT FOR FILING"])
    ws.append(["Generated", generated_at])
    ws.append(["Purpose", "Workpaper only. Use this to review the 2025 return with your CPA or tax software."])
    ws.append(["Status", "Not ready to file from this workbook alone. Open items remain."])
    ws.append(["Current date", "2026-03-13"])
    ws.append(["2025 Form 1120-S due date", "2026-03-16"])
    ws.append(["Extension", "File Form 7004 by 2026-03-16 if you need more time."])
    ws.append(["Cash-basis ordinary income (books)", money(float(page_1_lines["22"]["cash"]))])
    ws.append(["Cash-basis ordinary income (after $25k expense adj)", money(float(page_1_lines["22"]["cash"]) - personal_card_expense)])
    ws.append(["Accrual-basis ordinary income (books)", money(float(page_1_lines["22"]["accrual"]))])
    ws.append(["Accrual-basis ordinary income (after $25k expense adj)", money(float(page_1_lines["22"]["accrual"]) - personal_card_expense)])
    ws.append(["Shareholder loan repayment", loan_repayment])
    ws.append(["Net matched shareholder withdrawals", shareholder_match.get("Net withdrawals matched to 6611/0273 transfer family", "")])
    ws.append(["Unaccounted withdrawals after loan + $25k expense", shareholder_match.get("Unaccounted withdrawals after both items", "")])
    ws.append(["IRS Form 1120-S", "https://www.irs.gov/forms-pubs/about-form-1120-s"])
    ws.append(["IRS Instructions for Form 1120-S (2025)", "https://www.irs.gov/instructions/i1120s"])
    ws.append(["IRS Form 7004", "https://www.irs.gov/forms-pubs/about-form-7004"])
    ws.append(["IRS Form 3115", "https://www.irs.gov/forms-pubs/about-form-3115"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = warning_fill
    for row in ws.iter_rows(min_row=8, max_row=13, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '#,##0.00;[Red]-#,##0.00'

    ws2 = wb.create_sheet("Entity Header")
    ws2.append(["Field", "Value", "Source", "Note"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in ENTITY_INFO:
        ws2.append(list(row))

    ws3 = wb.create_sheet("Page1 Draft")
    ws3.freeze_panes = "A2"
    ws3.append(
        [
            "Form Line",
            "Label",
            "Cash Books",
            "Cash + $25k Adj",
            "Accrual Books",
            "Accrual + $25k Adj",
            "Status",
            "Note",
        ]
    )
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ordered_lines = sorted(page_1_lines.items(), key=lambda item: PAGE_1_LINE_ORDER.get(item[0], 999))
    for line_code, row in ordered_lines:
        cash_value = float(row["cash"])
        accrual_value = float(row["accrual"])
        cash_adj = cash_value + personal_card_expense if line_code in {"20", "21"} else cash_value - personal_card_expense if line_code == "22" else cash_value
        accrual_adj = accrual_value + personal_card_expense if line_code in {"20", "21"} else accrual_value - personal_card_expense if line_code == "22" else accrual_value
        ws3.append(
            [
                line_code,
                row["label"],
                money(cash_value),
                money(cash_adj),
                money(accrual_value),
                money(accrual_adj),
                "Draft value from current packet",
                row["note"],
            ]
        )

    for line_code, label in EXTRA_PAGE_1_LINES:
        note = "Leave blank until tax deposits, penalties, and refund/credit instructions are confirmed."
        if line_code == "23a":
            note = "No excess passive income or LIFO recapture tax was identified in the current packet, but this still needs tax-preparer confirmation."
        elif line_code == "23b":
            note = "No built-in gains Schedule D tax was identified in the current packet."
        elif line_code == "23c":
            note = "No additional tax or penalty form was identified in the current packet."
        ws3.append([line_code, label, None, None, None, None, "Blank pending support", note])

    for row in ws3.iter_rows(min_row=2, min_col=3, max_col=6):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00;[Red]-#,##0.00'

    ws4 = wb.create_sheet("Schedule K Draft")
    ws4.freeze_panes = "A2"
    ws4.append(
        [
            "Schedule K Line",
            "Label",
            "Cash Books",
            "Cash + $25k Adj",
            "Accrual Books",
            "Accrual + $25k Adj",
            "Status",
            "Note",
        ]
    )
    for cell in ws4[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in build_schedule_k_rows(page_1_lines, loan_repayment, personal_card_expense):
        ws4.append(row)
    for row in ws4.iter_rows(min_row=2, min_col=3, max_col=6):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00;[Red]-#,##0.00'

    ws5 = wb.create_sheet("Shareholder Draft")
    ws5.append(["Field", "Value", "Status", "Note"])
    for cell in ws5[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    shareholder_rows = [
        ("Shareholder name", "Bertrand A. Byrne", "Assumed from 2024 K-1", "Verify ownership remained unchanged in 2025."),
        ("Shareholder ID", "Omitted in this draft workbook", "Intentional", "Pull from original tax records or tax software."),
        ("Shareholder type", "Individual", "Assumed from 2024 K-1", ""),
        ("Ownership percentage", "100.00%", "Assumed from 2024 K-1", "Verify no ownership changes in 2025."),
        ("Cash-basis ordinary business income (loss)", money(float(page_1_lines["22"]["cash"])), "Draft value", "Would flow through if filing on cash with no extra $25k adjustment."),
        ("Cash-basis ordinary business income (loss) after $25k adj", money(float(page_1_lines["22"]["cash"]) - personal_card_expense), "Draft value", "Assumes the personal-card business expense is deductible in 2025 and not already booked."),
        ("Accrual-basis ordinary business income", money(float(page_1_lines["22"]["accrual"])), "Draft value", "Would flow through if filing on accrual with no extra $25k adjustment."),
        ("Accrual-basis ordinary business income after $25k adj", money(float(page_1_lines["22"]["accrual"]) - personal_card_expense), "Draft value", "Assumes the personal-card business expense is deductible in 2025 and not already booked."),
        ("Repayment of loans from shareholder", loan_repayment, "Manual adjustment", "Track for Schedule K informational reporting and shareholder basis review."),
        ("Distributions", "", "Unresolved", "Do not auto-fill distributions while owner activity is still under review."),
    ]
    for row in shareholder_rows:
        ws5.append(list(row))
    for row in ws5.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00;[Red]-#,##0.00'

    ws6 = wb.create_sheet("Open Items")
    ws6.append(["Type", "Item"])
    for cell in ws6[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for item in missing_items:
        ws6.append(["Missing", item])
    for item in review_items:
        ws6.append(["Review", item])

    ws7 = wb.create_sheet("Shareholder Support")
    ws7.append(["Metric", "Amount", "Source"])
    for cell in ws7[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for key, value in shareholder_match.items():
        ws7.append([key, value, "2025-tax-shareholder-loan-match.csv"])

    for sheet in wb.worksheets:
        autosize_columns(sheet)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)

    summary_lines = [
        "# 2025 Draft Form 1120-S Packet",
        "",
        "Status: DRAFT NOT FOR FILING",
        "",
        f"Generated: {generated_at}",
        "",
        "This draft workbook prepopulates the defensible 2025 Form 1120-S values currently supported by the AquatechPM packet.",
        "It does not fill fields that are still unsupported, and it should not be signed or filed as-is.",
        "",
        "Included workbook sheets:",
        "- Read Me",
        "- Entity Header",
        "- Page1 Draft",
        "- Schedule K Draft",
        "- Shareholder Draft",
        "- Open Items",
        "- Shareholder Support",
        "",
        "Key amounts:",
        f"- Cash-basis ordinary business income (books): {money(float(page_1_lines['22']['cash'])):,.2f}",
        f"- Cash-basis ordinary business income after $25,000 personal-card adjustment: {money(float(page_1_lines['22']['cash']) - personal_card_expense):,.2f}",
        f"- Accrual-basis ordinary business income (books): {money(float(page_1_lines['22']['accrual'])):,.2f}",
        f"- Accrual-basis ordinary business income after $25,000 personal-card adjustment: {money(float(page_1_lines['22']['accrual']) - personal_card_expense):,.2f}",
        f"- Shareholder loan repayment carried as a manual adjustment: {loan_repayment:,.2f}",
        "",
        "Do not finalize the filing until the open items on the `Open Items` sheet are resolved, especially the 2024 accounting method, depreciation, payroll filings, shareholder basis/Form 7203, and receipts/accountable-plan support.",
    ]
    OUTPUT_MD.write_text("\n".join(summary_lines) + "\n", encoding="utf-8")

    copied_paths: list[Path] = []
    for path in [OUTPUT_XLSX, OUTPUT_MD]:
        for destination_dir in [DOWNLOADS, TAX_FOLDER]:
            destination_dir.mkdir(parents=True, exist_ok=True)
            destination = destination_dir / path.name
            shutil.copy2(path, destination)
            copied_paths.append(destination)

    return [OUTPUT_XLSX, OUTPUT_MD, *copied_paths]


def main() -> None:
    for path in build_outputs():
        print(f"Wrote: {path}")


if __name__ == "__main__":
    main()
