"""
Timesheet generation from the REAL employee timekeeping workbooks.

Each employee maintains `Employee_Time_Keeping/.../<Name> ... TIMESHEETS.xlsx` with one
pre-formatted sheet per week ("WEEK NN 2026"), already filled with their logged hours.
The invoicing step just adds the print setup, the blue box around the billed (LTCP4)
row, weekend shading, and signatures — then exports that sheet to PDF.

We NEVER modify the live workbook: it's copied to a temp file first.

Layout (columns B..L): B=PROJECT#, C=TITLE, D=TASK#, E..K = Mon..Sun, L=TASK TOTAL.
Rows: 13 header, 14.. project rows (LTCP4 = row 15), then TOTAL DAILY / WEEKLY TOTAL.
"""
from __future__ import annotations
import datetime as dt, glob, os, shutil, tempfile
import openpyxl
from openpyxl.styles import PatternFill, Border, Side

TK_ROOT = r"G:\Shared drives\Aquatech Engineering PC\Employee_Time_Keeping"
WEEKEND_FILL = "FAE2D5"
BILLED_TITLE = "LTCP4"
DATA_FIRST, DATA_LAST = 14, 31       # project rows
DAY_COLS = list("EFGHIJK")           # Mon..Sun
WEEKEND_COLS = ["J", "K"]            # Sat, Sun

# AqtPM project name -> timesheet project code (column B). ALL projects an employee
# works appear on the one timesheet; only the invoice's project is outlined.
AQTPM_TO_CODE = {
    "LTCP4": "0020012024",
    "Aquatech Operations": "100004", "No Project": "100004",   # -> ADMINISTRATION
    "BWT Design Assistance": "0040042025",                     # Stantec JV
    "Brentwood Brook": "3",
    "Mount Vernon Flood Study": "0020032025",
    "Hydraulic Modeling 4063001X": "0010012024",              # BEPA
    "Business Development": "100005",
    "BWT 1608-Jobcon": "10052026",                            # JobCon -> row 19 slot
}
# projects that repurpose an empty numbered slot: code -> (row, code_to_set, title_to_set)
SPECIAL_ROWS = {"10052026": (19, 10052026, "BWT 1608-Jobcon")}


def _norm(code):
    if isinstance(code, float) and code == int(code):
        return str(int(code))
    return str(code).strip()


def workbook_for(full_name: str) -> str | None:
    hits = glob.glob(os.path.join(TK_ROOT, "*", f"{full_name}*TIMESHEETS.xlsx"))
    hits = [h for h in hits if "Archive" not in h]
    return hits[0] if hits else None


def _find_week_sheet(wb, week_ending: dt.date):
    for ws in wb.worksheets:
        v = ws["C8"].value
        if isinstance(v, dt.datetime) and v.date() == week_ending:
            return ws
        if isinstance(v, dt.date) and v == week_ending:
            return ws
    return None


def _billed_row(ws) -> int | None:
    for r in range(DATA_FIRST, DATA_LAST + 1):
        if str(ws[f"C{r}"].value or "").strip().upper() == BILLED_TITLE.upper():
            return r
    return None


def build_from_workbook(full_name: str, week_ending: dt.date, out_xlsx: str, *,
                        emp_signature: str | None = None,
                        sup_signature: str | None = None,
                        project_hours: dict | None = None) -> dict | None:
    """project_hours: {aqtpm_project_name: {date: hours}} for ALL projects the employee
    worked that week, from AqtPM/FreshBooks. Every project is shown on the one timesheet;
    only the invoice's project (LTCP4) is outlined. Fills FreshBooks-only loggers and
    keeps timesheet hours consistent with the source of truth."""
    src = workbook_for(full_name)
    if not src:
        return None
    tmp = os.path.join(tempfile.gettempdir(), f"_tk_{full_name.replace(' ','_')}.xlsx")
    shutil.copyfile(src, tmp)                     # never touch the live workbook

    # 1) read computed values (week-ending + day headers + totals are formulas that
    #    reference an EXTERNAL workbook, so we must bake their cached values).
    wb_vals = openpyxl.load_workbook(tmp, data_only=True)
    vs = _find_week_sheet(wb_vals, week_ending)
    if vs is None:
        return None
    title = vs.title
    values = {c.coordinate: c.value for row in vs.iter_rows() for c in row
              if c.value is not None}

    # 2) editable workbook; bake formulas -> values on the target sheet so it's
    #    self-contained (no external link / inter-sheet deps after we drop other sheets).
    wb = openpyxl.load_workbook(tmp)
    ws = wb[title]
    # Bake ONLY the external-workbook formulas (week-ending C8 etc.) to their cached
    # values. Keep in-sheet SUM formulas (task/daily/weekly totals) as formulas so they
    # recompute after we fill hours.
    for row in ws.iter_rows():
        for c in row:
            # bake any formula that references another sheet (internal or external,
            # marked by "!") — those break when we drop the other sheets. Keep pure
            # in-sheet SUM formulas so task/daily/weekly totals recompute after fills.
            if c.data_type == "f" and isinstance(c.value, str) and "!" in c.value:
                c.value = values.get(c.coordinate)
    for other in [s for s in wb.worksheets if s is not ws]:
        wb.remove(other)
    # drop the external-workbook link (its formulas are now baked to values)
    try:
        wb._external_links = []
    except Exception:
        pass

    # force the employee name from the known identity — some workbook sheets were copied
    # from another employee and carry the wrong typed name (e.g. Robert's showing "Stacey").
    _parts = full_name.split()
    ws["C5"] = _parts[0]
    ws["D5"] = " ".join(_parts[1:]) if len(_parts) > 1 else ""

    # weekend shading on the project rows
    fill = PatternFill("solid", fgColor=WEEKEND_FILL)
    for r in range(DATA_FIRST, DATA_LAST + 1):
        for col in WEEKEND_COLS:
            ws[f"{col}{r}"].fill = fill

    # fill ALL project rows from AqtPM/FreshBooks (every project the employee worked).
    if project_hours:
        code_row = {}
        for r in range(DATA_FIRST, DATA_LAST + 1):
            b = ws[f"B{r}"].value
            if b is not None:
                code_row[_norm(b)] = r
        # clear all project-row hours first so the total = AqtPM total exactly
        for r in range(DATA_FIRST, DATA_LAST + 1):
            for col in DAY_COLS:
                ws[f"{col}{r}"] = None
        for proj, daily in project_hours.items():
            code = AQTPM_TO_CODE.get(proj)
            if not code:
                continue
            if code in SPECIAL_ROWS:                 # repurpose a numbered slot (JobCon)
                row, set_code, set_title = SPECIAL_ROWS[code]
                ws[f"B{row}"] = set_code; ws[f"C{row}"] = set_title
            else:
                row = code_row.get(_norm(code))
            if not row:
                continue
            for j, col in enumerate(DAY_COLS):
                d = week_ending - dt.timedelta(days=6 - j)
                h = daily.get(d)
                if h:                                # accumulate (e.g. Ops + No Project)
                    cur = ws[f"{col}{row}"].value or 0
                    ws[f"{col}{row}"] = cur + h

    # blue box around the billed (LTCP4) row — the line item for THIS invoice
    br = _billed_row(ws)
    if br:
        blue = Side(style="medium", color="0000FF")
        cols = ["B", "C", "D"] + DAY_COLS + ["L"]
        for i, col in enumerate(cols):
            c = ws[f"{col}{br}"]; b = c.border
            c.border = Border(left=blue if i == 0 else b.left,
                              right=blue if i == len(cols) - 1 else b.right,
                              top=blue, bottom=blue)

    # signatures
    _stamp_signatures(ws, emp_signature, sup_signature)

    # print setup: landscape, fit 1 page, print area
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = "B2:L35"     # include a buffer row so WEEKLY TOTAL (34) isn't clipped
    ws.page_margins.left = ws.page_margins.right = 0.25
    ws.page_margins.top = ws.page_margins.bottom = 0.25
    ws.page_setup.horizontalCentered = True

    wb.save(out_xlsx)
    week_total = ws["E34"].value if isinstance(ws["E34"].value, (int, float)) else None
    return {"out": out_xlsx, "source": src, "sheet": ws.title,
            "billed_row": br, "week_total": week_total}


def _stamp_signatures(ws, emp_signature, sup_signature):
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    from PIL import Image as PILImage

    def place_centered(path, col, row, box_w, box_h, target_h):
        """Fit the signature (aspect preserved) into a box_w x box_h region anchored at
        (col,row) and CENTER it within that region."""
        im = PILImage.open(path)
        asp = im.width / im.height
        h = min(target_h, box_h)
        w = h * asp
        if w > box_w:                      # too wide -> shrink to fit
            w = box_w; h = w / asp
        coloff = max(0, (box_w - w) / 2)
        rowoff = max(0, (box_h - h) / 2)
        img = XLImage(path)
        img.anchor = OneCellAnchor(
            _from=AnchorMarker(col=col, colOff=pixels_to_EMU(coloff),
                               row=row, rowOff=pixels_to_EMU(rowoff)),
            ext=XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h)))
        ws.add_image(img)

    # employee e-signature: centered in the e-SIGNATURE row (row 6), below the typed
    # name so it doesn't overlap it
    if emp_signature and os.path.exists(emp_signature):
        place_centered(emp_signature, col=2, row=5, box_w=200, box_h=26, target_h=24)
    # supervisor approval: centered in the C-D region, rows 10-11
    if sup_signature and os.path.exists(sup_signature):
        place_centered(sup_signature, col=2, row=10, box_w=200, box_h=34, target_h=28)


if __name__ == "__main__":
    import io, sys
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    import config
    from timesheet_xlsx import ExcelSession
    r = build_from_workbook("Zachary Gilliam", dt.date(2026, 6, 7),
                            os.path.join(config.ROOT, "_wbts.xlsx"),
                            emp_signature=config.employee_signature("Zachary Gilliam"),
                            sup_signature=config.SUPERVISOR_SIG)
    print("result:", r)
    if r:
        with ExcelSession() as s:
            s.export(r["out"], os.path.join(config.ROOT, "_wbts.pdf"))
        print("exported _wbts.pdf")
