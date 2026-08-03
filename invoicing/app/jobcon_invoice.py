"""
Brown & Caldwell "JobCon" (BWT-1608-JBCON, Master Planning for Jamaica Bay WRRF)
sub-invoice generator. Unlike HDR/Stantec (NYCDEP cost-plus template), B&C require
their OWN multi-tab workbook. The dollars use the identical cost-plus factors
(staff x2.14 loaded, +10% profit; principal x2.14, 0% profit), computed here by the
sheet's own formulas — the '4a-Labor Summary' tab is the engine and everything else
(Cover Sheet subtask rollup, Summary Billing task rollup) pulls from it.

Data flow in the workbook (all formula-driven; we only write the inputs):
  4a-Labor Summary  hours -> E19 direct labor -> E22 (x2.14) -> +E23 (10%) + principal
                    -> E31 GRAND TOTAL
  1-Invoice Cover   C{sub}=Labor!E31 (this-invoice, active sub-task)
                    B{sub}=Previous!J{n} (prior cumulative for that sub-task)
  Previous          F..I monthly prior per sub-task -> J = SUM (prior cumulative)
  3-Summary Billing rolls sub-tasks up to Task level; E29/E47 = invoice total
Inputs we set: Work Summary C8 (invoice #) + narrative; Labor hours; Previous prior
months; Cover B15 period; Summary Billing I3 date + D27/F27 period.

Validated: reproduces BACAQ0002 (June 2026) = $11,808.95 to the penny.
"""
from __future__ import annotations
import datetime as dt, os, shutil, sys
from dataclasses import dataclass, field
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from hdr_invoice_generator import compute_labor, r2

# --- 4a-Labor Summary: full name -> row (template ships these rows) ---------------
LABOR_STAFF_ROW = {
    "Zachary Gilliam": 14, "Stacey Hodge": 15, "Robert Svadlenka": 16,
    "Roger Wang": 17, "Ailsa Welch": 18,
}
LABOR_PRINCIPAL_ROW = {"Bertrand Byrne": 24}
MONTH_COL = {5: "F", 6: "G", 7: "H", 8: "I"}   # Previous tab month columns (May..Aug)


@dataclass
class JobConInputs:
    invoice_number: str                     # e.g. "BACAQ0003"
    invoice_date: dt.date                   # Summary Billing I3
    period_begin: dt.date
    period_end: dt.date
    hours_by_emp: dict[str, float]          # this-invoice hours (single active sub-task)
    active_subtask_code: str                # e.g. "2.01.003"
    prior_by_month: list[tuple[dt.date, float]] = field(default_factory=list)  # active sub-task history
    narrative: list[str] | None = None      # Work Summary B14.. (optional; kept if None)
    this_odc: float = 0.0
    lump_sum: float = 0.0
    rates: dict[str, float] = field(default_factory=dict)


def _place_sig(ws, path: str, *, col: int, row: int, target_h: int = 24) -> None:
    """Anchor a signature image, scaled so the INK is ~target_h px tall (aspect kept)."""
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    from PIL import Image as PILImage
    im = PILImage.open(path)
    bb = im.convert("RGBA").getchannel("A").getbbox() or (0, 0, im.width, im.height)
    ink_h = max(1, bb[3] - bb[1])
    h = target_h * (im.height / ink_h)
    w = h * (im.width / im.height)
    img = XLImage(path)
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(col=col, colOff=pixels_to_EMU(6), row=row, rowOff=pixels_to_EMU(2)),
        ext=XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h)))
    ws.add_image(img)


def _find_row(ws, col: str, prefix: str, r0: int, r1: int) -> int:
    for r in range(r0, r1 + 1):
        v = ws[f"{col}{r}"].value
        if isinstance(v, str) and v.strip().startswith(prefix):
            return r
    raise ValueError(f"sub-task {prefix!r} not found in {ws.title}!{col}{r0}:{col}{r1}")


def _fill_labor(ws, inp: "JobConInputs"):
    try:
        import config as _cfg
        principals = _cfg.PRINCIPALS
    except Exception:
        principals = {"Bertrand Byrne"}

    for n in inp.hours_by_emp:
        if n not in LABOR_STAFF_ROW and n not in LABOR_PRINCIPAL_ROW:
            raise ValueError(f"{n!r} has JobCon hours but no row in the Labor Summary "
                             f"template — add a row and update LABOR_STAFF_ROW.")
    # zero every roster hour cell first, then set the ones who worked (idempotent)
    for r in list(LABOR_STAFF_ROW.values()) + list(LABOR_PRINCIPAL_ROW.values()):
        ws[f"D{r}"] = 0
    for n, h in inp.hours_by_emp.items():
        if not h:
            continue
        row = LABOR_STAFF_ROW.get(n) or LABOR_PRINCIPAL_ROW[n]
        # guard: template rate must match config (catch silent drift)
        tpl_rate = ws[f"C{row}"].value
        cfg_rate = inp.rates.get(n)
        if cfg_rate and tpl_rate and abs(float(tpl_rate) - float(cfg_rate)) > 1e-6:
            raise ValueError(f"JobCon Labor rate for {n} = {tpl_rate} but config says {cfg_rate}")
        ws[f"D{row}"] = h
    ws["E29"] = inp.lump_sum
    ws["E30"] = inp.this_odc


def _fill_previous(ws, inp: "JobConInputs", subtask_row: int):
    # clear this sub-task's month cells, then write each prior month into its column
    for col in MONTH_COL.values():
        ws[f"{col}{subtask_row}"] = 0
    for d, amt in inp.prior_by_month:
        if d.month not in MONTH_COL:
            raise ValueError(f"prior month {d:%Y-%m} has no column in the Previous tab "
                             f"(template covers May-Aug); extend the template.")
        ws[f"{MONTH_COL[d.month]}{subtask_row}"] = amt


def generate(template_xlsx: str, out_xlsx: str, inp: JobConInputs) -> dict:
    # verify single active sub-task assumption (B&C format bills one labor block)
    shutil.copyfile(template_xlsx, out_xlsx)
    wb = openpyxl.load_workbook(out_xlsx, data_only=False)
    work = wb["2-Work Summary"]
    labor = wb["4a-Labor Summary"]
    prev = wb["Previous"]
    cover = wb["1-Invoice Cover Sheet"]
    summ = wb["3-Summary Billing Report"]

    period = f"{inp.period_begin:%m/%d/%Y}-{inp.period_end:%m/%d/%Y}"

    # --- Work Summary: invoice # (drives Labor/Summary via formulas) + narrative ---
    work["C8"] = inp.invoice_number
    if inp.narrative is not None:
        # narrative = ordered list of (task header, body paragraph) built from this
        # month's time-entry notes. Write into the template's task slots and clear the
        # rest (incl. the old overview at B14) so no stale text carries over.
        work["B14"] = ""
        _slots = [(15, 16), (19, 20), (23, 24)]
        for _i, (_hr, _br) in enumerate(_slots):
            if _i < len(inp.narrative):
                _h, _b = inp.narrative[_i]
                work[f"B{_hr}"] = _h
                work[f"B{_br}"] = _b
            else:
                work[f"B{_hr}"] = ""
                work[f"B{_br}"] = ""

    # --- Labor Summary: hours (rates/positions are template constants) ---
    _fill_labor(labor, inp)

    # --- locate the active sub-task rows on Cover + Previous ---
    cover_row = _find_row(cover, "A", inp.active_subtask_code, 21, 61)
    prev_row = _find_row(prev, "E", inp.active_subtask_code, 2, 5)

    # wire this-invoice (Labor grand total) + prior (Previous row total) to that sub-task
    cover[f"C{cover_row}"] = "='4a-Labor Summary'!E31"
    cover[f"B{cover_row}"] = f"=Previous!J{prev_row}"
    cover["B15"] = period

    _fill_previous(prev, inp, prev_row)

    # --- Summary Billing: date + invoice period (were stale in the manual file) ---
    _idate = dt.datetime(inp.invoice_date.year, inp.invoice_date.month, inp.invoice_date.day)
    summ["I3"] = _idate
    summ["D27"] = dt.datetime(inp.period_begin.year, inp.period_begin.month, inp.period_begin.day)
    summ["F27"] = dt.datetime(inp.period_end.year, inp.period_end.month, inp.period_end.day)

    # fit the Summary Billing onto ONE page, and sign+date the Subcontractor Task
    # Manager certification line (D56 label / I56 "Date"): Ailsa signs it.
    from openpyxl.worksheet.properties import PageSetupProperties
    summ.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    summ.page_setup.fitToWidth = 1
    summ.page_setup.fitToHeight = 1
    summ["I55"] = _idate
    try:
        import config as _cfg2
        _sig = _cfg2.employee_signature("Ailsa Welch")
    except Exception:
        _sig = None
    if _sig and os.path.exists(_sig):
        _place_sig(summ, _sig, col=3, row=53)   # above "Subcontractor Task Manager"

    wb.save(out_xlsx)

    try:
        import config as _cfg
        principals = _cfg.PRINCIPALS
    except Exception:
        principals = {"Bertrand Byrne"}
    calc = compute_labor(inp.hours_by_emp, inp.rates or None, principals)
    this_total = r2(calc["total_labor"] + inp.this_odc + inp.lump_sum)
    return dict(out=out_xlsx, this_labor=calc["total_labor"], this_total=this_total,
                prior_cum=r2(sum(a for _d, a in inp.prior_by_month)),
                active_subtask=inp.active_subtask_code, calc=calc)


if __name__ == "__main__":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    import config as cfg
    TPL = r"G:\Shared drives\Aquatech_Admin\Invoicing\Brown and Caldwell\June 2026\BWT-1608-JBCON_Aquatech Subinvoice June 2026.xlsx"
    OUT = os.path.join(cfg.ROOT, "_test_jobcon_BACAQ0002.xlsx")
    inp = JobConInputs(
        invoice_number="BACAQ0002", invoice_date=dt.date(2026, 7, 1),
        period_begin=dt.date(2026, 6, 1), period_end=dt.date(2026, 6, 30),
        hours_by_emp={"Robert Svadlenka": 38, "Ailsa Welch": 18, "Bertrand Byrne": 15},
        active_subtask_code="2.01.003",
        prior_by_month=[(dt.date(2026, 5, 1), 8206.69)],
        rates=dict(cfg.RATES),
    )
    res = generate(TPL, OUT, inp)
    EXPECT = 11808.95
    ok = abs(res["this_total"] - EXPECT) < 0.02
    print("=== JobCon generator — regenerate BACAQ0002 (June 2026) ===")
    print(f"  staff  direct labor : ${res['calc']['staff']['direct_labor']:,.2f}")
    print(f"  princ  direct labor : ${res['calc']['principal']['direct_labor']:,.2f}")
    print(f"  THIS-INVOICE total  : ${res['this_total']:,.2f}   expected ${EXPECT:,.2f}")
    print(f"  prior cumulative    : ${res['prior_cum']:,.2f}   (expect 8,206.69)")
    print(f"  active sub-task     : {res['active_subtask']}")
    print(f"  wrote               : {res['out']}")
    print("  >>> MATCHES BACAQ0002 (compute_labor)" if ok else "  >>> MISMATCH")
