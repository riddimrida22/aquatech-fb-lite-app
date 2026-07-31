"""
HDR / LTCP4 sub-consultant invoice generator.

Fills the real HDR NYCDEP cost-plus template xlsx from AqtPM hours, for a chosen
LTCP4 billing period. The template is almost entirely formula-driven, so generation
only patches a handful of cells and rolls the `Prior` (cumulative) sheet cleanly.

Validated: regenerating the May 2026 invoice reproduces HDRAQ0017 = $25,052.21 to the penny.

Cells patched (sheet -> cell):
  Invoice Summary  G11 = invoice date (mm/dd/YYYY)
  Invoice Summary  G12 = invoice number (HDRAQ####)
  Invoice Summary  C20 = period begin (datetime)
  Invoice Summary  E20 = period end (datetime)
  Invoice Summary  C29 = prior Other-Direct-Costs cumulative (number)
  Invoice Summary  C28 = =Prior!B{total_row}   (prior labor cumulative, range-correct)
  Task# 5          H14 = Zachary hours   L14 = rate
  Task# 5          H15 = Stacey hours    L15 = rate
  Task# 5          H16 = Robert hours    L16 = rate
  Task# 5          H22 = Bertrand hours  L22 = rate (principal, profit 0)
  Prior            one clean row per historical invoice + Total = SUM(all rows)

Everything else (OH 114%, profit 10% staff / 0% principal, ROUND(x,2) per component,
task rollup, % complete, balance remaining) recomputes from the existing formulas.
"""
from __future__ import annotations
import csv, datetime as dt, os, shutil
from dataclasses import dataclass, field
import openpyxl

# ---- reference data --------------------------------------------------------
OH_RATE = 1.14
PROFIT_RATE = 0.10

# Approved BILLING rates (fixed per person; actual salary is higher). Principal bills no profit.
RATES_2026 = {
    "Zachary Gilliam": 52.26,
    "Stacey Hodge":    51.64,
    "Robert Svadlenka":61.50,
    "Bertrand Byrne":  97.50,
    "Ailsa Welch":     75.00,
}
PRINCIPALS = {"Bertrand Byrne"}

# Row map in the "Invoice Template_Task# 5" sheet (staff block + principal block).
STAFF_ROWS = {"Zachary Gilliam": 14, "Stacey Hodge": 15, "Robert Svadlenka": 16}
PRINCIPAL_ROW = {"Bertrand Byrne": 22}

CONTRACT = dict(
    bill_to_name="Henningson, Durham & Richardson Architecture and Engineering, P.C.",
    contract_id="CSO-LTCP04: Combined Sewer Overflow Long Term Control Plan-04",
    po="PO# 1000100113624",
    labor_max=1277886.72,
    odc_max=8200.0,
)


def _here() -> str:
    return os.path.dirname(os.path.abspath(__file__))


def load_billing_periods(path: str | None = None) -> dict[int, tuple[dt.date, dt.date]]:
    path = path or os.path.join(_here(), "ltcp4_billing_periods_2026.csv")
    out: dict[int, tuple[dt.date, dt.date]] = {}
    with open(path, newline="") as f:
        for row in csv.DictReader(f):
            b = dt.date.fromisoformat(row["begin"])
            e = dt.date.fromisoformat(row["end"])
            out[int(row["period"])] = (b, e)
    return out


def r2(x: float) -> float:
    return round(x + 1e-9, 2)


# ---- cost-plus math (mirror of the sheet formulas, for preview/validation) --
def compute_labor(hours_by_emp: dict[str, float],
                  rates: dict[str, float] | None = None,
                  principals: set[str] | None = None) -> dict:
    """Return the same numbers the sheet will compute, for preview + validation.
    rates/principals default to this module's constants but can be supplied (e.g. the
    invoicing app's config, which carries every active employee)."""
    rates = rates or RATES_2026
    principals = principals if principals is not None else PRINCIPALS
    staff = {n: h for n, h in hours_by_emp.items() if n not in principals}
    prin  = {n: h for n, h in hours_by_emp.items() if n in principals}

    def block(members: dict[str, float], apply_profit: bool):
        detail, labor = [], 0.0
        for n, h in members.items():
            if n not in rates:
                raise ValueError(f"No billing rate for {n!r}")
            rate = rates[n]; amt = r2(h * rate); labor += amt
            detail.append(dict(name=n, hours=h, rate=rate, amount=amt))
        labor = r2(labor)
        oh = r2(labor * OH_RATE)
        profit = r2((labor + oh) * PROFIT_RATE) if apply_profit else 0.0
        return dict(detail=detail, labor=labor, oh=oh, profit=profit,
                    direct_labor=r2(labor + oh + profit))

    s = block(staff, True)
    p = block(prin, False)
    return dict(staff=s, principal=p, total_labor=r2(s["direct_labor"] + p["direct_labor"]))


# ---- generator -------------------------------------------------------------
@dataclass
class PriorInvoice:
    date: dt.date          # first-of-month tag used in the Prior sheet
    labor_amount: float    # THIS-INVOICE labor for that historical invoice


@dataclass
class InvoiceInputs:
    period_number: int
    invoice_number: str            # e.g. "HDRAQ0017"
    invoice_date: dt.date
    hours_by_emp: dict[str, float]
    prior_invoices: list[PriorInvoice] = field(default_factory=list)
    prior_odc_cumulative: float = 0.0
    this_odc: float = 0.0
    rates: dict[str, float] = field(default_factory=lambda: dict(RATES_2026))
    authorized_sig: str | None = None       # path to signer's signature image
    authorized_by: str = "Bertrand Byrne, CEO"


def _emp_meta(name: str, inp: "InvoiceInputs"):
    """(first, surname, title, rate) for an employee, from config if available."""
    try:
        import config as _cfg
        title = _cfg.TITLES.get(name, "Engineer")
    except Exception:
        title = "Engineer"
    rate = inp.rates.get(name, RATES_2026.get(name))
    parts = name.split()
    first = parts[0]
    surname = " ".join(parts[1:]) if len(parts) > 1 else ""
    return first, surname, title, rate


def _copy_row_style(ws, src_row: int, dst_row: int, cols=range(1, 15)):
    from copy import copy
    for c in cols:
        s = ws.cell(row=src_row, column=c)
        d = ws.cell(row=dst_row, column=c)
        d.font = copy(s.font); d.border = copy(s.border)
        d.fill = copy(s.fill); d.alignment = copy(s.alignment)
        d.number_format = s.number_format


def _fill_task5(task, summ, inp: "InvoiceInputs"):
    """Fill the STAFF + PRINCIPAL cost-plus blocks for any headcount. The template ships
    with 3 staff rows (14-16) + 1 principal (22); extra people get inserted rows with the
    same formatting, and every downstream formula is rewritten at its new position."""
    try:
        import config as _cfg
        order = _cfg.STAFF_ORDER
        principals_set = _cfg.PRINCIPALS
    except Exception:
        order, principals_set = [], PRINCIPALS

    def _order_key(n):
        return (order.index(n) if n in order else len(order), n)

    staff = sorted([n for n, h in inp.hours_by_emp.items()
                    if n not in principals_set and h], key=_order_key)
    prin = sorted([n for n, h in inp.hours_by_emp.items()
                   if n in principals_set and h], key=_order_key)

    TPL_STAFF, TPL_PRIN = 3, 2   # template staff rows 14-16, principal slots 22-23
    extra_staff = max(0, len(staff) - TPL_STAFF)
    if extra_staff:
        task.insert_rows(17, extra_staff)              # before the staff SUM row
        for i in range(extra_staff):
            _copy_row_style(task, 16, 17 + i)

    sh = extra_staff                                   # downstream shift
    staff_sum = 17 + sh
    oh_s, prof_s, dir_s = 18 + sh, 19 + sh, 20 + sh
    prin_first = 22 + sh
    prin_sum = 24 + sh
    oh_p, prof_p, dir_p = 25 + sh, 26 + sh, 27 + sh
    total_labor = 29 + sh
    exp_val, exp_sum, task_total = 32 + sh, 34 + sh, 37 + sh

    # staff rows 14..
    for i in range(TPL_STAFF + extra_staff):
        r = 14 + i
        if i < len(staff):
            n = staff[i]; first, sur, title, rate = _emp_meta(n, inp)
            task[f"A{r}"] = first; task[f"B{r}"] = sur; task[f"D{r}"] = title
            task[f"H{r}"] = inp.hours_by_emp[n]; task[f"L{r}"] = rate
            task[f"N{r}"] = f"=ROUND(H{r}*L{r},2)"
        else:  # unused template staff row -> blank
            for col in "ABDH":
                task[f"{col}{r}"] = None
            task[f"L{r}"] = None; task[f"N{r}"] = None
    task[f"H{staff_sum}"] = f"=SUM(H13:H{staff_sum-1})"
    task[f"N{staff_sum}"] = f"=SUM(N13:N{staff_sum-1})"
    task[f"N{oh_s}"] = f"=ROUND(N{staff_sum}*1.14,2)"
    task[f"N{prof_s}"] = f"=ROUND((N{staff_sum}+N{oh_s})*0.1,2)"
    task[f"N{dir_s}"] = f"=SUM(N{staff_sum}:N{prof_s})"

    # principal rows (template has 2 slots 22-23; fill n_prin, blank the rest)
    for i in range(TPL_PRIN):
        r = prin_first + i
        if i < len(prin):
            n = prin[i]; first, sur, title, rate = _emp_meta(n, inp)
            task[f"A{r}"] = first; task[f"B{r}"] = sur; task[f"D{r}"] = title
            task[f"H{r}"] = inp.hours_by_emp[n]; task[f"L{r}"] = rate
            task[f"N{r}"] = f"=ROUND(H{r}*L{r},2)"
        else:
            for col in "ABDH":
                task[f"{col}{r}"] = None
            task[f"L{r}"] = None; task[f"N{r}"] = None
    task[f"H{prin_sum}"] = f"=SUM(H{prin_first}:H{prin_first+TPL_PRIN-1})"
    task[f"N{prin_sum}"] = f"=SUM(N{prin_first}:N{prin_first+TPL_PRIN-1})"
    task[f"N{oh_p}"] = f"=ROUND(N{prin_sum}*1.14,2)"
    task[f"N{prof_p}"] = f"=ROUND((N{prin_sum}+N{oh_p})*0,2)"
    task[f"N{dir_p}"] = f"=SUM(N{prin_sum}:N{prof_p})"
    task[f"N{total_labor}"] = f"=N{dir_p}+N{dir_s}"

    # The template HIDES the principal "Profit @ 0%" row (always zero). insert_rows does
    # not move row heights/hidden flags, so re-apply: show the OH row, hide profit@0.
    if oh_p in task.row_dimensions:
        task.row_dimensions[oh_p].hidden = False
    task.row_dimensions[prof_p].hidden = True

    # expenses
    task[f"N{exp_val}"] = f"=L{exp_val}"
    if inp.this_odc:
        task[f"L{exp_val}"] = inp.this_odc
    task[f"N{exp_sum}"] = f"=SUM(N{exp_val}:N{exp_val+1})"
    task[f"N{task_total}"] = f"=N{exp_sum}+N{total_labor}"

    # repoint the Invoice Summary at the (possibly shifted) Task# 5 cells
    T = "'Invoice Template_Task# 5'"
    summ["D27"] = f"={T}!N{task_total}"
    summ["D28"] = f"={T}!N{dir_s}+{T}!N{dir_p}"
    summ["D29"] = f"={T}!N{exp_sum}"


def generate(template_xlsx: str, out_xlsx: str, inp: InvoiceInputs,
             periods: dict[int, tuple[dt.date, dt.date]] | None = None) -> dict:
    periods = periods or load_billing_periods()
    begin, end = periods[inp.period_number]

    shutil.copyfile(template_xlsx, out_xlsx)
    wb = openpyxl.load_workbook(out_xlsx, data_only=False)
    summ = wb["Invoice Summary"]
    task = wb["Invoice Template_Task# 5"]
    prior = wb["Prior"]

    # --- Invoice Summary header ---
    summ["G11"] = inp.invoice_date.strftime("%m/%d/%Y")
    summ["G12"] = inp.invoice_number
    summ["G13"] = CONTRACT["po"]
    summ["C20"] = dt.datetime(begin.year, begin.month, begin.day)
    summ["E20"] = dt.datetime(end.year, end.month, end.day)
    summ["C29"] = inp.prior_odc_cumulative          # prior ODC cumulative
    summ["F28"] = CONTRACT["labor_max"]
    summ["F29"] = CONTRACT["odc_max"]

    # --- Task# 5: fill staff + principal blocks, DYNAMIC headcount ---
    _fill_task5(task, summ, inp)

    # --- Prior sheet: clean one-row-per-invoice + correct Total range ---
    # Clear any existing rows from row 5 down, then rewrite.
    for r in range(5, prior.max_row + 2):
        prior[f"A{r}"] = None
        prior[f"B{r}"] = None
    r = 5
    prior[f"A{r}"] = "Prior"
    if inp.prior_invoices:
        prior[f"B{r}"] = inp.prior_invoices[0].labor_amount
        for pi in inp.prior_invoices[1:]:
            r += 1
            prior[f"A{r}"] = dt.datetime(pi.date.year, pi.date.month, pi.date.day)
            prior[f"B{r}"] = pi.labor_amount
    else:
        prior[f"B{r}"] = 0
    last_data_row = r
    total_row = last_data_row + 1
    prior[f"A{total_row}"] = "Total"
    prior[f"B{total_row}"] = f"=SUM(B5:B{last_data_row})"
    # point Summary C28 at the correct Total cell
    summ["C28"] = f"=Prior!B{total_row}"

    # authorized signature on the Invoice Summary, above the "Authorized Signature" line
    # (F38), with the signer name/title beside it — matches the previous submitted invoice.
    if inp.authorized_sig and os.path.exists(inp.authorized_sig):
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU
        from PIL import Image as PILImage
        _im = PILImage.open(inp.authorized_sig); _h = 34; _w = _h * (_im.width / _im.height)
        _img = XLImage(inp.authorized_sig)
        _img.anchor = OneCellAnchor(
            _from=AnchorMarker(col=5, colOff=pixels_to_EMU(15), row=35, rowOff=pixels_to_EMU(2)),
            ext=XDRPositiveSize2D(pixels_to_EMU(_w), pixels_to_EMU(_h)))
        summ.add_image(_img)
        summ["H36"] = inp.authorized_by
        summ["H37"] = "Aquatech Engineering P.C."

    wb.save(out_xlsx)

    try:
        import config as _cfg
        _principals = _cfg.PRINCIPALS
    except Exception:
        _principals = PRINCIPALS
    calc = compute_labor(inp.hours_by_emp, inp.rates or None, _principals)
    return dict(out=out_xlsx, period=(begin.isoformat(), end.isoformat()),
                this_labor=calc["total_labor"], this_odc=inp.this_odc,
                this_total=r2(calc["total_labor"] + inp.this_odc),
                prior_labor_cum=r2(sum(pi.labor_amount for pi in inp.prior_invoices)),
                calc=calc)


if __name__ == "__main__":
    # VALIDATION: regenerate the May 2026 invoice (period 5) from the April template.
    import io, sys
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    HDR = r"G:\Shared drives\Aquatech_Admin\Invoicing\HDR"
    APRIL = os.path.join(HDR, "HDR Invoice April 2026",
                         "HDR Sub-Consultant Invoicing_Aquatech April 2026 05132026 .xlsx")
    OUT = os.path.join(_here(), "_generated_May2026_from_April.xlsx")

    # prior invoices = the historical labor amounts through April (from the real Prior sheet)
    prior_vals = [
        (None,           13838.31),  ("2025-04-01", 22235.67), ("2025-05-01", 26899.80),
        ("2025-06-01", 28065.03), ("2025-07-01", 39980.02), ("2025-08-01", 34147.98),
        ("2025-09-01", 22236.74), ("2025-10-01", 26507.11), ("2025-11-01", 33545.57),
        ("2025-12-01", 15664.80), ("2026-01-01", 28033.47), ("2026-02-01", 32212.17),
        ("2026-03-01", 51456.99), ("2026-04-01", 38955.55),
    ]
    priors = [PriorInvoice(dt.date(2000,1,1) if d is None else dt.date.fromisoformat(d), v)
              for d, v in prior_vals]

    inp = InvoiceInputs(
        period_number=5, invoice_number="HDRAQ0017", invoice_date=dt.date(2026, 6, 2),
        hours_by_emp={"Zachary Gilliam": 97, "Robert Svadlenka": 20, "Bertrand Byrne": 49},
        prior_invoices=priors, prior_odc_cumulative=5960.63, this_odc=0.0,
    )
    res = generate(APRIL, OUT, inp)
    EXPECT = 25052.21
    ok = abs(res["this_total"] - EXPECT) < 0.005
    print("=== HDR generator — regenerate May 2026 (period 5) from April template ===")
    print(f"  period            : {res['period'][0]} -> {res['period'][1]}")
    print(f"  staff  direct labor: ${res['calc']['staff']['direct_labor']:,.2f}")
    print(f"  princ  direct labor: ${res['calc']['principal']['direct_labor']:,.2f}")
    print(f"  THIS-INVOICE total : ${res['this_total']:,.2f}   expected ${EXPECT:,.2f}")
    print(f"  prior labor cum    : ${res['prior_labor_cum']:,.2f}")
    print(f"  wrote              : {res['out']}")
    print("  >>> PENNY-EXACT MATCH TO HDRAQ0017" if ok else "  >>> MISMATCH")
