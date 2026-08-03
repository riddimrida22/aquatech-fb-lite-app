"""
Pixel-perfect Aquatech Weekly Timesheet generator.

Reconstructs the exact "AQUATECH ENGINEERING P.C. WEEKLY TIMESHEET" form (landscape),
fills hours from AqtPM time_entries, and (on Windows w/ Excel) exports to PDF with the
same rendering as the real print-to-PDF.

Layout matched field-for-field to the real May-2026 HDR package timesheets:
 - bordered title bar, employee box (First Name / Surname split), WEEK ENDING box,
   SUPERVISOR box, "HOURS" banner
 - fixed 18-row project grid with project codes + task#, Mon..Sun day columns,
   weekend (Sat/Sun) columns shaded #FAE2D5, TASK TOTAL column
 - TOTAL DAILY HOURS row + WEEKLY TOTAL row
"""
from __future__ import annotations
import datetime as dt
import os
from dataclasses import dataclass
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WEEKEND_FILL = "FAE2D5"
HEADER_GRAY = "F3F3F3"
FONT = "Times New Roman"   # the real form uses a serif face

# Fixed project grid (code, title, task#, italic?) — exactly as on the real timesheet.
PROJECT_ROWS = [
    ("0010012024", "BEPA HYDRAULIC MODELING", "3", False),
    ("0020012024", "LTCP4", "1", False),
    ("3", "Brentwood Brook", "1", False),
    ("0040042025", "BWT Design Assistance (Stantec JV)", "2", False),
    ("0020032025", "Mount Vernon", "7", False),
    ("10052026", "BWT 1608-Jobcon", "2", False), ("7", "", "", False), ("8", "", "", False),
    ("9", "", "", False), ("10", "", "", False), ("11", "", "", False),
    ("12", "", "", False), ("13", "", "", False),
    ("100001", "PUBLIC HOLIDAY", "", True),
    ("100002", "SICK LEAVE", "", True),
    ("100003", "VACATION", "", True),
    ("100004", "ADMINISTRATION", "", True),
    ("100005", "BUSINESS DEVELOPMENT", "", True),
]

# AqtPM project name -> timesheet project code.
AQTPM_TO_CODE = {
    "LTCP4": "0020012024",
    "Aquatech Operations": "100004",       # -> ADMINISTRATION
    "No Project": "100004",
    "BWT Design Assistance": "0040042025",
    "Brentwood Brook": "3",
    "Mount Vernon Flood Study": "0020032025",
    "Hydraulic Modeling 4063001X": "0010012024",
    "Business Development": "100005",
    "BWT 1608-Jobcon": "10052026",
}
SUPERVISOR = "BERTRAND BYRNE, PhD, P.E"

thin = Side(style="thin", color="000000")
med = Side(style="medium", color="000000")
GRID = Border(left=thin, right=thin, top=thin, bottom=thin)


@dataclass
class DayEntry:
    project: str
    date: dt.date
    hours: float
    note: str = ""


def _box(ws, r1, c1, r2, c2, side=med):
    """Draw a border box around a cell range (outer edges only)."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            b = cell.border
            left = side if c == c1 else b.left
            right = side if c == c2 else b.right
            top = side if r == r1 else b.top
            bottom = side if r == r2 else b.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def build_timesheet_xlsx(out_path: str, *, employee_first: str, employee_last: str,
                         week_ending: dt.date, entries: list[DayEntry],
                         billed_project: str = "LTCP4",
                         emp_signature: str | None = None,
                         sup_signature: str | None = None) -> dict:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4

    # column widths matched to the real timesheet geometry (measured from the PDF:
    # PROJECT# 97pt, TITLE 116pt, TASK# 53pt, day cols ~46.5pt, TASK TOTAL 50pt).
    widths = [16, 25, 9, 8.3, 8.3, 8.3, 8.3, 8.3, 8.3, 8.3, 11]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    days = [week_ending - dt.timedelta(days=6 - i) for i in range(7)]  # Mon..Sun

    def setc(coord, val, *, bold=False, italic=False, size=9, center=True,
             fill=None, border=GRID, wrap=False):
        c = ws[coord]
        c.value = val
        c.font = Font(name=FONT, bold=bold, italic=italic, size=size)
        c.alignment = Alignment(horizontal="center" if center else "left",
                                vertical="center", wrap_text=wrap)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        if border:
            c.border = border
        return c

    # --- Title bar (row 1) ---
    ws.merge_cells("A1:K1")
    setc("A1", "AQUATECH ENGINEERING P.C. WEEKLY TIMESHEET", bold=True, size=14)
    _box(ws, 1, 1, 1, 11)
    ws.row_dimensions[1].height = 22

    # --- Employee box (rows 3-5) A3:H5 ---
    ws.row_dimensions[2].height = 6
    ws.merge_cells("D3:E3"); setc("D3", "First Name", bold=True, border=None)
    ws.merge_cells("F3:G3"); setc("F3", "Surname", bold=True, border=None)
    setc("A4", "EMPLOYEE NAME:", center=False, border=None)
    ws.merge_cells("D4:E4"); setc("D4", employee_first, border=None)
    ws.merge_cells("F4:G4"); setc("F4", employee_last, border=None)
    setc("A5", "EMPLOYEE e-SIGNATURE", center=False, border=None)
    _box(ws, 3, 1, 5, 8)

    # --- WEEK ENDING box (row 7) A7:F7 ---
    ws.row_dimensions[6].height = 6
    setc("A7", "WEEK ENDING", center=False, border=None)
    ws.merge_cells("C7:F7")
    we = f"{week_ending.strftime('%A')}, {week_ending.strftime('%B')} {week_ending.day}, {week_ending.year}"
    setc("C7", we, bold=True, center=False, border=None)
    _box(ws, 7, 1, 7, 6)

    # --- SUPERVISOR box (rows 9-10) A9:H10 ---
    ws.row_dimensions[8].height = 6
    setc("A9", "SUPERVISOR NAME:", center=False, border=None)
    ws.merge_cells("D9:F9"); setc("D9", SUPERVISOR, border=None)
    setc("A10", "SUPERVISOR APPROVAL", center=False, border=None)
    _box(ws, 9, 1, 10, 8)

    # --- HOURS banner (row 11 over day columns) ---
    ws.merge_cells("D11:J11")
    setc("D11", "HOURS", bold=True, border=None)

    # --- Grid header (row 12) ---
    hdr_row = 12
    headers = ["PROJECT#", "PROJECT TITLE", "TASK#"] + \
        [f"{d.month}/{d.day}/{str(d.year)[2:]}" for d in days] + ["TASK TOTAL"]
    for ci, h in enumerate(headers, start=1):
        setc(ws.cell(row=hdr_row, column=ci).coordinate, h, bold=True, fill=HEADER_GRAY)
    ws.row_dimensions[hdr_row].height = 18

    # --- aggregate AqtPM hours by code+date ---
    hours: dict[str, dict[dt.date, float]] = {}
    for e in entries:
        code = AQTPM_TO_CODE.get(e.project)
        if not code:
            continue
        hours.setdefault(code, {})
        hours[code][e.date] = hours[code].get(e.date, 0.0) + float(e.hours)

    # --- project rows ---
    first_data = hdr_row + 1
    daily_totals = {d: 0.0 for d in days}
    week_total = 0.0
    billed_code = AQTPM_TO_CODE.get(billed_project)
    billed_row = None
    for i, (code, title, task, italic) in enumerate(PROJECT_ROWS):
        r = first_data + i
        if code == billed_code:
            billed_row = r
        setc(f"A{r}", code, italic=italic)
        setc(f"B{r}", title, italic=italic, center=False)
        setc(f"C{r}", task)
        row_total = 0.0
        for di, d in enumerate(days):
            col = 4 + di  # D..J
            h = hours.get(code, {}).get(d, 0.0)
            weekend = d.weekday() >= 5  # Sat=5, Sun=6
            cell = setc(ws.cell(row=r, column=col).coordinate,
                        (h if h else None), fill=(WEEKEND_FILL if weekend else None))
            if h:
                daily_totals[d] += h
                row_total += h
        week_total += row_total
        setc(f"K{r}", row_total if row_total else 0)
        ws.row_dimensions[r].height = 15

    last_data = first_data + len(PROJECT_ROWS) - 1

    # --- enclose the billed (LTCP4) row in a blue box (matches the real timesheet) ---
    if billed_row is not None:
        blue = Side(style="medium", color="0000FF")
        for c in range(1, 12):  # A..K
            cell = ws.cell(row=billed_row, column=c)
            b = cell.border
            cell.border = Border(
                left=blue if c == 1 else b.left,
                right=blue if c == 11 else b.right,
                top=blue, bottom=blue)

    # --- TOTAL DAILY HOURS row ---
    tdr = last_data + 1
    ws.merge_cells(f"A{tdr}:C{tdr}")
    setc(f"A{tdr}", "TOTAL DAILY HOURS", bold=True)
    for di, d in enumerate(days):
        setc(ws.cell(row=tdr, column=4 + di).coordinate,
             (daily_totals[d] if daily_totals[d] else 0), bold=True)
    setc(f"K{tdr}", None)
    ws.row_dimensions[tdr].height = 18

    # --- WEEKLY TOTAL row ---
    wtr = tdr + 1
    ws.merge_cells(f"A{wtr}:C{wtr}")
    setc(f"A{wtr}", "WEEKLY TOTAL", bold=True)
    ws.merge_cells(f"D{wtr}:J{wtr}")
    setc(f"D{wtr}", week_total, bold=True)
    setc(f"K{wtr}", None)
    ws.row_dimensions[wtr].height = 18

    _box(ws, hdr_row, 1, wtr, 11)     # outer box around whole grid

    # --- signatures (employee e-signature + supervisor approval) ---
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU

    def _place(path, col, row, coloff, rowoff, w, h):
        # Scale so the INK is ~h tall and preserve aspect ratio (no stretch), so a
        # signature with transparent margin renders at the same visible size/shape as
        # the prior submitted timesheets instead of being squashed to a fixed box.
        import PIL.Image as _PILImage
        _im = _PILImage.open(path).convert("RGBA")
        _bb = _im.getchannel("A").getbbox() or (0, 0, _im.width, _im.height)
        _ink_h = max(1, _bb[3] - _bb[1])
        _H = h * (_im.height / _ink_h)
        _W = _H * (_im.width / _im.height)
        img = XLImage(path)
        marker = AnchorMarker(col=col, colOff=pixels_to_EMU(coloff),
                              row=row, rowOff=pixels_to_EMU(rowoff))
        img.anchor = OneCellAnchor(_from=marker,
                                   ext=XDRPositiveSize2D(pixels_to_EMU(_W), pixels_to_EMU(_H)))
        ws.add_image(img)

    # cols are 0-indexed: D=3, E=4; rows 0-indexed: r-1. Tuck signatures inside their boxes.
    if emp_signature and os.path.exists(emp_signature):
        _place(emp_signature, col=4, row=3, coloff=2, rowoff=2, w=128, h=30)   # E4 area
    if sup_signature and os.path.exists(sup_signature):
        _place(sup_signature, col=2, row=8, coloff=4, rowoff=2, w=120, h=28)    # C9 area

    wb.save(out_path)
    return dict(out=out_path, week_ending=str(week_ending), week_total=week_total,
                ltcp4_hours=sum(hours.get("0020012024", {}).values()),
                daily_totals={str(k): v for k, v in daily_totals.items()})


def _xml_unescape(s: str) -> str:
    return (s.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
             .replace("&quot;", '"').replace("&apos;", "'"))


def _prep_xlsx_for_pdf(src_xlsx: str, dst_xlsx: str, sheet_names: list[str] | None) -> None:
    """Copy src_xlsx -> dst_xlsx, editing ONLY xl/workbook.xml to (a) hide every sheet
    not in sheet_names and (b) force fullCalcOnLoad. Media/drawings/styles are copied
    verbatim so embedded images survive (openpyxl would drop them)."""
    import re, zipfile
    with zipfile.ZipFile(src_xlsx) as zin:
        wbxml = zin.read("xl/workbook.xml").decode("utf-8")

    # ordered list of sheet names as they appear in the workbook
    all_names = [_xml_unescape(m.group(1))
                 for m in re.finditer(r'<sheet\b[^>]*\bname="([^"]*)"', wbxml)]
    if sheet_names:
        keep = {n for n in sheet_names if n in all_names} or ({all_names[0]} if all_names else set())
    else:
        keep = {all_names[0]} if all_names else set()

    def _fix_sheet(m: "re.Match") -> str:
        tag = m.group(0)
        nm_m = re.search(r'\bname="([^"]*)"', tag)
        nm = _xml_unescape(nm_m.group(1)) if nm_m else ""
        if nm in keep:
            # ensure kept sheets are visible (a template may ship one hidden)
            if 'state=' in tag:
                tag = re.sub(r'\bstate="[^"]*"', 'state="visible"', tag)
            return tag
        if 'state=' in tag:
            return re.sub(r'\bstate="[^"]*"', 'state="hidden"', tag)
        # inject state="hidden" just before the self-closing />
        return re.sub(r'\s*/>\s*$', ' state="hidden"/>', tag)

    wbxml = re.sub(r'<sheet\b[^>]*/>', _fix_sheet, wbxml)

    # CRITICAL: LibreOffice's PDF export renders EVERY sheet that has a defined print
    # area, regardless of the hidden flag. So hiding alone doesn't isolate sheets — we
    # must also strip the Print_Area defined name for every non-kept sheet (referenced by
    # its 0-based localSheetId). Kept sheets keep their print areas.
    keep_idx = {i for i, nm in enumerate(all_names) if nm in keep}

    def _drop_print_area(m: "re.Match") -> str:
        dn = m.group(0)
        if 'Print_Area' not in dn:
            return dn
        lsid = re.search(r'localSheetId="(\d+)"', dn)
        if lsid and int(lsid.group(1)) not in keep_idx:
            return ""      # this sheet isn't being exported — remove its print range
        return dn

    wbxml = re.sub(r'<definedName\b[^>]*>.*?</definedName>', _drop_print_area, wbxml, flags=re.S)

    # force a full recalc on open (openpyxl formulas have no cached values)
    if 'fullCalcOnLoad' not in wbxml:
        if '<calcPr' in wbxml:
            wbxml = re.sub(r'<calcPr\b', '<calcPr fullCalcOnLoad="1" ', wbxml, count=1)
        else:
            wbxml = wbxml.replace('</workbook>', '<calcPr fullCalcOnLoad="1"/></workbook>')

    with zipfile.ZipFile(src_xlsx) as zin, \
         zipfile.ZipFile(dst_xlsx, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = wbxml.encode("utf-8") if item.filename == "xl/workbook.xml" else zin.read(item.filename)
            zout.writestr(item, data)


class ExcelSession:
    """Reusable session for exporting xlsx sheets to PDF, platform-aware:
      - Windows: drives Excel via COM (pixel-perfect; the local tool).
      - Linux / no Excel (the cloud server): uses LibreOffice headless.
    Same .export(xlsx, pdf, sheet_names) API either way, so callers don't change.
    """
    # Excel's ExportAsFixedFormat needs a WORKING default printer to compute page
    # layout; if the OS default is a physical/offline printer (e.g. a Canon) it HANGS.
    # So on Windows we force "Microsoft Print to PDF" for the session and restore after.
    PDF_PRINTER = "Microsoft Print to PDF"

    def __init__(self):
        self.mode = "libreoffice"
        self.excel = None
        self._orig_printer = None
        try:
            import win32com.client.dynamic as dyn  # noqa: F401  (Windows only)
            self.mode = "excel"
        except Exception:
            self.mode = "libreoffice"
        if self.mode == "excel":
            import win32com.client.dynamic as dyn
            try:
                import win32print
                self._orig_printer = win32print.GetDefaultPrinter()
                if self._orig_printer != self.PDF_PRINTER:
                    win32print.SetDefaultPrinter(self.PDF_PRINTER)
            except Exception:
                self._orig_printer = None  # best-effort; export still tries
            self.excel = dyn.Dispatch("Excel.Application")
            self.excel.Visible = False
            self.excel.DisplayAlerts = False

    def export(self, xlsx_path: str, pdf_path: str, sheet_names: list[str] | None = None) -> str:
        if self.mode == "excel":
            return self._export_excel(xlsx_path, pdf_path, sheet_names)
        return self._export_libreoffice(xlsx_path, pdf_path, sheet_names)

    def _export_excel(self, xlsx_path: str, pdf_path: str, sheet_names: list[str] | None = None) -> str:
        import os
        wb = self.excel.Workbooks.Open(os.path.abspath(xlsx_path))
        try:
            if sheet_names:
                wb.Worksheets(sheet_names[0]).Select()
                for nm in sheet_names[1:]:
                    wb.Worksheets(nm).Select(False)
                self.excel.ActiveSheet.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
            else:
                wb.Sheets(1).ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        finally:
            wb.Close(False)
        return pdf_path

    def _export_libreoffice(self, xlsx_path: str, pdf_path: str, sheet_names: list[str] | None = None) -> str:
        """Render the requested sheet(s) to PDF with headless LibreOffice.

        We select sheets by HIDING the others at the zip/XML level (edit xl/workbook.xml
        only) instead of re-saving through openpyxl. openpyxl drops embedded images on a
        load+save round-trip, which would strip the Aquatech/HDR logo and the signatures
        off the invoice and timesheet PDFs. Zip surgery leaves xl/media and xl/drawings
        untouched, so the output is byte-for-byte the same artwork Excel would print.
        LibreOffice excludes hidden sheets from a PDF export, and fullCalcOnLoad forces it
        to recompute the openpyxl-written formulas (which carry no cached values)."""
        import os, subprocess, tempfile, shutil, glob
        tmpdir = tempfile.mkdtemp(prefix="loexp_")
        try:
            src = os.path.join(tmpdir, "in.xlsx")
            _prep_xlsx_for_pdf(xlsx_path, src, sheet_names)
            prof = "file://" + os.path.join(tmpdir, "profile")
            subprocess.run(
                ["soffice", "--headless", "-env:UserInstallation=" + prof,
                 "--calc", "--convert-to", "pdf", "--outdir", tmpdir, src],
                check=True, capture_output=True, timeout=180)
            out = os.path.join(tmpdir, "in.pdf")
            if not os.path.exists(out):
                found = glob.glob(os.path.join(tmpdir, "*.pdf"))
                if not found:
                    raise RuntimeError("LibreOffice produced no PDF for " + xlsx_path)
                out = found[0]
            dest_dir = os.path.dirname(os.path.abspath(pdf_path)) or "."
            os.makedirs(dest_dir, exist_ok=True)
            shutil.move(out, pdf_path)
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)
        return pdf_path

    def close(self):
        if self.mode == "excel" and self.excel is not None:
            try:
                self.excel.Quit()
            except Exception:
                pass
            try:
                import win32print
                if self._orig_printer and self._orig_printer != self.PDF_PRINTER:
                    win32print.SetDefaultPrinter(self._orig_printer)
            except Exception:
                pass

    def __enter__(self):
        return self

    def __exit__(self, *a):
        self.close()


def export_pdf_via_excel(xlsx_path: str, pdf_path: str) -> str:
    """Export the xlsx to PDF using Excel COM (Windows). Pixel-identical to print-to-PDF."""
    with ExcelSession() as s:
        return s.export(xlsx_path, pdf_path)


if __name__ == "__main__":
    import io, os, sys
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    here = os.path.dirname(os.path.abspath(__file__))
    entries = [
        DayEntry("Aquatech Operations", dt.date(2026, 5, 4), 3),
        DayEntry("LTCP4", dt.date(2026, 5, 6), 6),
        DayEntry("LTCP4", dt.date(2026, 5, 7), 8),
        DayEntry("LTCP4", dt.date(2026, 5, 8), 8),
    ]
    xlsx = os.path.join(here, "_ts_zach_wk1.xlsx")
    res = build_timesheet_xlsx(xlsx, employee_first="Zachary", employee_last="Gilliam",
                               week_ending=dt.date(2026, 5, 10), entries=entries)
    print("weekly total:", res["week_total"], "LTCP4:", res["ltcp4_hours"])
    print("daily:", res["daily_totals"])
    try:
        pdf = os.path.join(here, "_ts_zach_wk1.pdf")
        export_pdf_via_excel(xlsx, pdf)
        print("PDF:", pdf)
    except Exception as e:
        print("PDF export skipped:", e)
